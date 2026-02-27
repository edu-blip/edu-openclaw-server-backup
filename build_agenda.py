#!/usr/bin/env python3
"""
Generates the Mussali family March agenda as an Excel file.
Upload to Google Drive → open as Google Sheets → share link.
"""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from datetime import date, timedelta

wb = Workbook()

# ─── Color palette ────────────────────────────────────────────────────────────
BLUE_HEADER   = "1F4E79"   # dark blue  – header background
BLUE_MID      = "2E75B6"   # mid blue   – sub-header
LIGHT_BLUE    = "BDD7EE"   # light blue – weekend / N/A cells
YELLOW_FILL   = "FFF2CC"   # yellow     – baby bath
GREEN_FILL    = "E2EFDA"   # green      – home visits
ORANGE_FILL   = "FCE4D6"   # orange     – school tasks
GRAY_FILL     = "D9D9D9"   # gray       – N/A cells
WHITE         = "FFFFFF"
INPUT_FILL    = "FFFEF0"   # very light yellow – editable cells

DAYS_ES = ["Lunes","Martes","Miércoles","Jueves","Viernes","Sábado","Domingo"]

def h_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border_thin():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def header_font(color="FFFFFF", bold=True, size=10):
    return Font(name="Calibri", bold=bold, color=color, size=size)

def cell_font(bold=False, size=10):
    return Font(name="Calibri", bold=bold, size=size)

def apply_header(cell, text, bg=BLUE_HEADER, fg="FFFFFF", size=11, bold=True):
    cell.value = text
    cell.fill = h_fill(bg)
    cell.font = header_font(color=fg, bold=bold, size=size)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border_thin()

def apply_label(cell, text, bg=ORANGE_FILL, bold=False, color="000000"):
    cell.value = text
    cell.fill = h_fill(bg)
    cell.font = Font(name="Calibri", bold=bold, size=9, color=color)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = border_thin()

def apply_input(cell, bg=INPUT_FILL):
    cell.fill = h_fill(bg)
    cell.font = Font(name="Calibri", size=10, color="1A1A1A")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border_thin()

def apply_na(cell):
    cell.value = "—"
    cell.fill = h_fill(GRAY_FILL)
    cell.font = Font(name="Calibri", size=10, color="AAAAAA")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border_thin()

def march_dates():
    start = date(2026, 3, 1)
    return [start + timedelta(days=i) for i in range(31)]

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 – Escuela de Emilio
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "📚 Escuela de Emilio"
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "A4"

# Title
ws1.merge_cells("A1:F1")
apply_header(ws1["A1"], "🏫  ESCUELA DE EMILIO — MARZO 2026",
             bg=BLUE_HEADER, size=14)
ws1.row_dimensions[1].height = 30

# Sub-header instructions
ws1.merge_cells("A2:F2")
ws1["A2"].value = (
    "Escribe tu nombre en la celda que quieras reservar. "
    "Primero en llegar, primero en reservar. "
    "Si ya tiene nombre, elige otro día."
)
ws1["A2"].fill = h_fill("EBF3FB")
ws1["A2"].font = Font(name="Calibri", italic=True, size=9, color="2E4057")
ws1["A2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws1["A2"].border = border_thin()
ws1.row_dimensions[2].height = 22

# Column headers
headers = ["Fecha", "Día", "Llevar a Chabad\n9:00am\n(Mar-Vie)",
           "Recoger Chabad→Village\n11:45am\n(Mar-Vie)",
           "Recoger de Village\n3:00pm\n(Lun-Vie)", "Notas"]
for col, h in enumerate(headers, 1):
    apply_header(ws1.cell(3, col), h, bg=BLUE_MID, size=9)
ws1.row_dimensions[3].height = 48

# Column widths
ws1.column_dimensions["A"].width = 12
ws1.column_dimensions["B"].width = 12
ws1.column_dimensions["C"].width = 20
ws1.column_dimensions["D"].width = 22
ws1.column_dimensions["E"].width = 20
ws1.column_dimensions["F"].width = 18

row = 4
for d in march_dates():
    dow = d.weekday()  # 0=Mon … 6=Sun
    day_name = DAYS_ES[dow]
    is_weekend = dow >= 5  # Sat or Sun
    is_tue_fri = 1 <= dow <= 4
    is_mon_fri = 0 <= dow <= 4

    # Date cell
    date_cell = ws1.cell(row, 1, value=d.strftime("%-d %b"))
    date_cell.fill = h_fill(LIGHT_BLUE if is_weekend else "DDEEFF")
    date_cell.font = Font(name="Calibri", bold=True, size=10,
                          color="5A5A5A" if is_weekend else "1F4E79")
    date_cell.alignment = Alignment(horizontal="center", vertical="center")
    date_cell.border = border_thin()

    # Day name cell
    day_cell = ws1.cell(row, 2, value=day_name)
    day_cell.fill = h_fill(LIGHT_BLUE if is_weekend else "DDEEFF")
    day_cell.font = Font(name="Calibri", bold=not is_weekend, size=10,
                         color="5A5A5A" if is_weekend else "1F4E79")
    day_cell.alignment = Alignment(horizontal="center", vertical="center")
    day_cell.border = border_thin()

    # Chabad drop-off (Tue-Fri)
    c3 = ws1.cell(row, 3)
    if is_tue_fri:
        apply_input(c3, bg=ORANGE_FILL)
    else:
        apply_na(c3)

    # Chabad→Village pickup (Tue-Fri)
    c4 = ws1.cell(row, 4)
    if is_tue_fri:
        apply_input(c4, bg=ORANGE_FILL)
    else:
        apply_na(c4)

    # Village pickup (Mon-Fri)
    c5 = ws1.cell(row, 5)
    if is_mon_fri:
        apply_input(c5, bg=ORANGE_FILL)
    else:
        apply_na(c5)

    # Notes
    c6 = ws1.cell(row, 6)
    apply_input(c6, bg=WHITE)

    ws1.row_dimensions[row].height = 22
    row += 1

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 – Baño del Bebé
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("🛁 Baño del Bebé")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "A4"

ws2.merge_cells("A1:D1")
apply_header(ws2["A1"], "🛁  BAÑO DEL BEBÉ — MARZO 2026",
             bg="7B2D8B", size=14)
ws2.row_dimensions[1].height = 30

ws2.merge_cells("A2:D2")
ws2["A2"].value = (
    "Solo 1 persona por día puede hacer el baño del bebé. "
    "Escribe tu nombre para reservar. Una vez reservado, ya no se puede cambiar sin avisar a Edu."
)
ws2["A2"].fill = h_fill("F3E5F5")
ws2["A2"].font = Font(name="Calibri", italic=True, size=9, color="4A1060")
ws2["A2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws2["A2"].border = border_thin()
ws2.row_dimensions[2].height = 22

for col, h in enumerate(["Fecha", "Día", "Tu Nombre 👶", "Notas"], 1):
    apply_header(ws2.cell(3, col), h, bg="9B59B6", size=10)
ws2.row_dimensions[3].height = 28

ws2.column_dimensions["A"].width = 12
ws2.column_dimensions["B"].width = 12
ws2.column_dimensions["C"].width = 28
ws2.column_dimensions["D"].width = 22

row = 4
for d in march_dates():
    dow = d.weekday()
    day_name = DAYS_ES[dow]
    is_weekend = dow >= 5

    c1 = ws2.cell(row, 1, value=d.strftime("%-d %b"))
    c1.fill = h_fill("E8D5F0" if is_weekend else "F3E5F5")
    c1.font = Font(name="Calibri", bold=True, size=10, color="7B2D8B")
    c1.alignment = Alignment(horizontal="center", vertical="center")
    c1.border = border_thin()

    c2 = ws2.cell(row, 2, value=day_name)
    c2.fill = h_fill("E8D5F0" if is_weekend else "F3E5F5")
    c2.font = Font(name="Calibri", bold=not is_weekend, size=10, color="7B2D8B")
    c2.alignment = Alignment(horizontal="center", vertical="center")
    c2.border = border_thin()

    c3 = ws2.cell(row, 3)
    apply_input(c3, bg=YELLOW_FILL)

    c4 = ws2.cell(row, 4)
    apply_input(c4, bg=WHITE)

    ws2.row_dimensions[row].height = 22
    row += 1

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 – Visitas a Casa
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("🏠 Visitas a Casa")
ws3.sheet_view.showGridLines = False
ws3.freeze_panes = "A4"

ws3.merge_cells("A1:E1")
apply_header(ws3["A1"], "🏠  VISITAS A CASA — MARZO 2026  (3:00pm – 5:00pm)",
             bg="1A6B3C", size=14)
ws3.row_dimensions[1].height = 30

ws3.merge_cells("A2:E2")
ws3["A2"].value = (
    "Puedes visitar cualquier día de 3:00pm a 5:00pm. "
    "Pueden venir varias familias el mismo día (hasta 3). "
    "Escribe tu nombre en uno de los espacios del día que elijas."
)
ws3["A2"].fill = h_fill("D5EDDF")
ws3["A2"].font = Font(name="Calibri", italic=True, size=9, color="1A4A2A")
ws3["A2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws3["A2"].border = border_thin()
ws3.row_dimensions[2].height = 22

for col, h in enumerate(["Fecha", "Día", "Visita 1 🧑", "Visita 2 🧑", "Visita 3 🧑"], 1):
    apply_header(ws3.cell(3, col), h, bg="27AE60", size=10)
ws3.row_dimensions[3].height = 28

ws3.column_dimensions["A"].width = 12
ws3.column_dimensions["B"].width = 12
ws3.column_dimensions["C"].width = 22
ws3.column_dimensions["D"].width = 22
ws3.column_dimensions["E"].width = 22

row = 4
for d in march_dates():
    dow = d.weekday()
    day_name = DAYS_ES[dow]
    is_weekend = dow >= 5

    c1 = ws3.cell(row, 1, value=d.strftime("%-d %b"))
    c1.fill = h_fill("C8E6D4" if is_weekend else "E8F5EE")
    c1.font = Font(name="Calibri", bold=True, size=10, color="1A6B3C")
    c1.alignment = Alignment(horizontal="center", vertical="center")
    c1.border = border_thin()

    c2 = ws3.cell(row, 2, value=day_name)
    c2.fill = h_fill("C8E6D4" if is_weekend else "E8F5EE")
    c2.font = Font(name="Calibri", bold=not is_weekend, size=10, color="1A6B3C")
    c2.alignment = Alignment(horizontal="center", vertical="center")
    c2.border = border_thin()

    for col in [3, 4, 5]:
        apply_input(ws3.cell(row, col), bg=GREEN_FILL)

    ws3.row_dimensions[row].height = 22
    row += 1

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 – Instrucciones
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("📋 Instrucciones")
ws4.sheet_view.showGridLines = False

ws4.column_dimensions["A"].width = 5
ws4.column_dimensions["B"].width = 70

ws4.merge_cells("A1:B1")
apply_header(ws4["A1"], "📋  INSTRUCCIONES — CÓMO USAR ESTA AGENDA",
             bg=BLUE_HEADER, size=14)
ws4.row_dimensions[1].height = 32

instructions = [
    ("", ""),
    ("🎉", "¡Bienvenidos a la agenda familiar de los Mussali para Marzo 2026!"),
    ("", "La bebé está a punto de llegar y queremos organizarnos para que todo fluya bien."),
    ("", "Esta agenda tiene 3 secciones. Cada una es una pestaña en la parte de abajo:"),
    ("", ""),
    ("📚", "PESTAÑA 1 — Escuela de Emilio"),
    ("", "  • Llevar a Chabad: martes a viernes, antes de las 9:00am"),
    ("", "  • Recoger de Chabad y llevar a Village: martes a viernes, ~11:45am"),
    ("", "  • Recoger de Village: lunes a viernes, ~3:00pm"),
    ("", "  ⚠️  Emilio debe estar en casa a las 5:30pm para su baño."),
    ("", ""),
    ("🛁", "PESTAÑA 2 — Baño del Bebé"),
    ("", "  • Una sola persona por día puede ayudar con el baño de la bebé."),
    ("", "  • Escribe tu nombre en el día que quieras reservar."),
    ("", ""),
    ("🏠", "PESTAÑA 3 — Visitas a Casa"),
    ("", "  • Las visitas son de 3:00pm a 5:00pm cualquier día."),
    ("", "  • Pueden venir hasta 3 familias el mismo día."),
    ("", "  • Escribe tu nombre en uno de los 3 espacios del día que elijas."),
    ("", ""),
    ("✅", "CÓMO RESERVAR"),
    ("", "  1. Abre la pestaña correspondiente (arriba)."),
    ("", "  2. Busca el día que te funcione."),
    ("", "  3. Haz clic en la celda vacía y escribe tu nombre."),
    ("", "  4. Guarda (Ctrl+S o Cmd+S). ¡Listo!"),
    ("", ""),
    ("❌", "REGLAS IMPORTANTES"),
    ("", "  • Si la celda ya tiene nombre, esa fecha está reservada. Elige otra."),
    ("", "  • No borres el nombre de otra persona."),
    ("", "  • Si necesitas cancelar, avisa a Edu directamente por WhatsApp."),
    ("", ""),
    ("💛", "¡Gracias por su apoyo! Con amor, Edu & familia 🍼"),
]

for i, (icon, text) in enumerate(instructions, 2):
    ws4.row_dimensions[i].height = 18
    if icon:
        cell_icon = ws4.cell(i, 1, value=icon)
        cell_icon.font = Font(name="Segoe UI Emoji", size=11)
        cell_icon.alignment = Alignment(horizontal="center", vertical="center")
    cell_text = ws4.cell(i, 2, value=text)
    if icon in ["📚", "🛁", "🏠", "✅", "❌", "💛"]:
        cell_text.font = Font(name="Calibri", bold=True, size=11, color="1F4E79")
        cell_text.fill = h_fill("EBF3FB")
    elif icon == "🎉":
        cell_text.font = Font(name="Calibri", bold=True, size=12, color="1F4E79")
    else:
        cell_text.font = Font(name="Calibri", size=10, color="2E2E2E")
    cell_text.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ── Save ──────────────────────────────────────────────────────────────────────
out = "/root/.openclaw/workspace/Agenda_Mussali_Marzo2026.xlsx"
wb.save(out)
print(f"✅ Saved: {out}")
