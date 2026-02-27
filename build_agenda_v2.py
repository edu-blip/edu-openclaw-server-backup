#!/usr/bin/env python3
"""
Mussali Family March Agenda — v2
Changes from v1:
- Added "Llevar a Village" column for Mondays (12pm)
- Dropdown data validation for all name cells (via a hidden Names sheet)
- Improved column layout & mobile-friendlier labels
- Tab names clarified for iPhone
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import date, timedelta

wb = Workbook()

# ─── Color palette ────────────────────────────────────────────────────────────
BLUE_HEADER = "1F4E79"
BLUE_MID    = "2E75B6"
LIGHT_BLUE  = "BDD7EE"
YELLOW_FILL = "FFF2CC"
GREEN_FILL  = "E2EFDA"
ORANGE_FILL = "FCE4D6"
PURPLE_BG   = "7B2D8B"
PURPLE_LIGHT= "F3E5F5"
PURPLE_MID  = "9B59B6"
GRAY_FILL   = "D9D9D9"
WHITE       = "FFFFFF"
INPUT_FILL  = "FFFEF0"

DAYS_ES = ["Lunes","Martes","Miércoles","Jueves","Viernes","Sábado","Domingo"]

NAMES = [
    "Beto y Tere",
    "Emilio y Miriam",
    "Chema y Monica",
    "Helen",
    "Sammy y Alice",
    "Eduardo y Esther",
    "Adolfo",
    "Amigos",
]

def h_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_header(cell, text, bg=BLUE_HEADER, fg="FFFFFF", size=11, bold=True):
    cell.value = text
    cell.fill = h_fill(bg)
    cell.font = Font(name="Calibri", bold=bold, color=fg, size=size)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()

def apply_input(cell, bg=INPUT_FILL):
    cell.fill = h_fill(bg)
    cell.font = Font(name="Calibri", size=10, color="1A1A1A")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border()

def apply_na(cell):
    cell.value = "—"
    cell.fill = h_fill(GRAY_FILL)
    cell.font = Font(name="Calibri", size=10, color="AAAAAA")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border()

def date_cell_style(cell, d, is_weekend, bg_normal, bg_weekend, color_normal, color_weekend):
    cell.fill = h_fill(bg_weekend if is_weekend else bg_normal)
    cell.font = Font(name="Calibri", bold=True, size=10,
                     color=color_weekend if is_weekend else color_normal)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border()

def march_dates():
    start = date(2026, 3, 1)
    return [start + timedelta(days=i) for i in range(31)]

# ══════════════════════════════════════════════════════════════════════════════
# HIDDEN SHEET: Nombres (for dropdown source)
# ══════════════════════════════════════════════════════════════════════════════
ws_names = wb.active
ws_names.title = "Nombres"
ws_names.sheet_state = "hidden"

apply_header(ws_names["A1"], "Lista de Nombres", bg=BLUE_HEADER)
for i, name in enumerate(NAMES, 2):
    ws_names.cell(i, 1, value=name)
    ws_names.cell(i, 1).font = Font(name="Calibri", size=10)
    ws_names.cell(i, 1).alignment = Alignment(horizontal="left")

ws_names.column_dimensions["A"].width = 25

# Named range string for dropdown — Excel-style absolute reference
names_range = f"Nombres!$A$2:$A${1 + len(NAMES)}"

def make_dv(sheet):
    """Create a dropdown DataValidation pointing to the Nombres sheet."""
    dv = DataValidation(
        type="list",
        formula1=names_range,
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Nombre no válido",
        error="Por favor elige un nombre de la lista.",
        showInputMessage=True,
        promptTitle="Selecciona tu nombre",
        prompt="Elige tu nombre del menú desplegable ▼"
    )
    sheet.add_data_validation(dv)
    return dv

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 – Escuela de Emilio
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("1 - Emilio 📚")
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "A4"

# Title
ws1.merge_cells("A1:F1")
apply_header(ws1["A1"], "🏫  ESCUELA DE EMILIO — MARZO 2026", bg=BLUE_HEADER, size=13)
ws1.row_dimensions[1].height = 30

# Subtitle
ws1.merge_cells("A2:F2")
ws1["A2"].value = "Elige tu nombre del menú ▼ en el día que puedas ayudar. Primero en llegar, primero en reservar."
ws1["A2"].fill = h_fill("EBF3FB")
ws1["A2"].font = Font(name="Calibri", italic=True, size=9, color="2E4057")
ws1["A2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws1["A2"].border = thin_border()
ws1.row_dimensions[2].height = 22

# Column headers
col_headers = [
    "Fecha",
    "Día",
    "Llevar a\nChabad\n⏰ 9:00am\n(Mar–Vie)",
    "Recoger Chabad\n→ llevar a Village\n⏰ 11:45am\n(Mar–Vie)",
    "Llevar a Village\n⏰ 12:00pm\n(Solo Lunes)",
    "Recoger de\nVillage\n⏰ 3:00pm\n(Lun–Vie)",
]
col_bgs = [BLUE_MID, BLUE_MID, BLUE_MID, BLUE_MID, "1A6B3C", BLUE_MID]
for col, (h, bg) in enumerate(zip(col_headers, col_bgs), 1):
    apply_header(ws1.cell(3, col), h, bg=bg, size=9)
ws1.row_dimensions[3].height = 56

ws1.column_dimensions["A"].width = 11
ws1.column_dimensions["B"].width = 12
ws1.column_dimensions["C"].width = 18
ws1.column_dimensions["D"].width = 22
ws1.column_dimensions["E"].width = 18
ws1.column_dimensions["F"].width = 18

dv1 = make_dv(ws1)

row = 4
input_cells_ws1 = []
for d in march_dates():
    dow = d.weekday()  # 0=Mon … 6=Sun
    day_name = DAYS_ES[dow]
    is_weekend = dow >= 5
    is_mon     = dow == 0
    is_tue_fri = 1 <= dow <= 4
    is_mon_fri = 0 <= dow <= 4

    # Date
    c1 = ws1.cell(row, 1, value=d.strftime("%-d %b"))
    date_cell_style(c1, d, is_weekend, "DDEEFF", LIGHT_BLUE, "1F4E79", "5A5A5A")
    # Day
    c2 = ws1.cell(row, 2, value=day_name)
    date_cell_style(c2, d, is_weekend, "DDEEFF", LIGHT_BLUE,
                    "1F4E79", "5A5A5A")
    c2.font = Font(name="Calibri", bold=not is_weekend, size=10,
                   color="5A5A5A" if is_weekend else "1F4E79")

    # Col C – Llevar a Chabad (Tue-Fri)
    c3 = ws1.cell(row, 3)
    if is_tue_fri: apply_input(c3, bg=ORANGE_FILL); input_cells_ws1.append(c3)
    else: apply_na(c3)

    # Col D – Recoger Chabad→Village (Tue-Fri)
    c4 = ws1.cell(row, 4)
    if is_tue_fri: apply_input(c4, bg=ORANGE_FILL); input_cells_ws1.append(c4)
    else: apply_na(c4)

    # Col E – Llevar a Village (Monday only)
    c5 = ws1.cell(row, 5)
    if is_mon: apply_input(c5, bg="D5EDDF"); input_cells_ws1.append(c5)
    else: apply_na(c5)

    # Col F – Recoger de Village (Mon-Fri)
    c6 = ws1.cell(row, 6)
    if is_mon_fri: apply_input(c6, bg=ORANGE_FILL); input_cells_ws1.append(c6)
    else: apply_na(c6)

    ws1.row_dimensions[row].height = 22
    row += 1

# Add dropdown validation to all input cells
for c in input_cells_ws1:
    dv1.add(c)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 – Baño del Bebé
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("2 - Baño Bebé 🛁")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "A4"

ws2.merge_cells("A1:D1")
apply_header(ws2["A1"], "🛁  BAÑO DE LA BEBÉ — MARZO 2026", bg=PURPLE_BG, size=13)
ws2.row_dimensions[1].height = 30

ws2.merge_cells("A2:D2")
ws2["A2"].value = "Solo 1 persona por día. Elige tu nombre del menú ▼. Una vez reservado, solo Edu puede cambiarlo."
ws2["A2"].fill = h_fill(PURPLE_LIGHT)
ws2["A2"].font = Font(name="Calibri", italic=True, size=9, color="4A1060")
ws2["A2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws2["A2"].border = thin_border()
ws2.row_dimensions[2].height = 22

for col, h in enumerate(["Fecha", "Día", "👶 Quién da el baño", "Notas"], 1):
    apply_header(ws2.cell(3, col), h, bg=PURPLE_MID, size=10)
ws2.row_dimensions[3].height = 28

ws2.column_dimensions["A"].width = 11
ws2.column_dimensions["B"].width = 12
ws2.column_dimensions["C"].width = 28
ws2.column_dimensions["D"].width = 22

dv2 = make_dv(ws2)
input_cells_ws2 = []

row = 4
for d in march_dates():
    dow = d.weekday()
    is_weekend = dow >= 5
    day_name = DAYS_ES[dow]

    c1 = ws2.cell(row, 1, value=d.strftime("%-d %b"))
    date_cell_style(c1, d, is_weekend, "F3E5F5", "E8D5F0", "7B2D8B", "7B2D8B")
    c2 = ws2.cell(row, 2, value=day_name)
    date_cell_style(c2, d, is_weekend, "F3E5F5", "E8D5F0", "7B2D8B", "7B2D8B")
    c2.font = Font(name="Calibri", bold=not is_weekend, size=10, color="7B2D8B")

    c3 = ws2.cell(row, 3)
    apply_input(c3, bg=YELLOW_FILL)
    input_cells_ws2.append(c3)

    c4 = ws2.cell(row, 4)
    apply_input(c4, bg=WHITE)

    ws2.row_dimensions[row].height = 22
    row += 1

for c in input_cells_ws2:
    dv2.add(c)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 – Visitas a Casa
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("3 - Visitas 🏠")
ws3.sheet_view.showGridLines = False
ws3.freeze_panes = "A4"

ws3.merge_cells("A1:E1")
apply_header(ws3["A1"], "🏠  VISITAS A CASA — MARZO 2026  (3:00pm – 5:00pm)", bg="1A6B3C", size=13)
ws3.row_dimensions[1].height = 30

ws3.merge_cells("A2:E2")
ws3["A2"].value = "Pueden venir hasta 3 grupos el mismo día, de 3:00pm a 5:00pm. Elige tu nombre del menú ▼."
ws3["A2"].fill = h_fill("D5EDDF")
ws3["A2"].font = Font(name="Calibri", italic=True, size=9, color="1A4A2A")
ws3["A2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws3["A2"].border = thin_border()
ws3.row_dimensions[2].height = 22

for col, h in enumerate(["Fecha", "Día", "Visita 1 🧑", "Visita 2 🧑", "Visita 3 🧑"], 1):
    apply_header(ws3.cell(3, col), h, bg="27AE60", size=10)
ws3.row_dimensions[3].height = 28

ws3.column_dimensions["A"].width = 11
ws3.column_dimensions["B"].width = 12
ws3.column_dimensions["C"].width = 22
ws3.column_dimensions["D"].width = 22
ws3.column_dimensions["E"].width = 22

dv3 = make_dv(ws3)
input_cells_ws3 = []

row = 4
for d in march_dates():
    dow = d.weekday()
    is_weekend = dow >= 5
    day_name = DAYS_ES[dow]

    c1 = ws3.cell(row, 1, value=d.strftime("%-d %b"))
    date_cell_style(c1, d, is_weekend, "E8F5EE", "C8E6D4", "1A6B3C", "1A6B3C")
    c2 = ws3.cell(row, 2, value=day_name)
    date_cell_style(c2, d, is_weekend, "E8F5EE", "C8E6D4", "1A6B3C", "1A6B3C")
    c2.font = Font(name="Calibri", bold=not is_weekend, size=10, color="1A6B3C")

    for col in [3, 4, 5]:
        c = ws3.cell(row, col)
        apply_input(c, bg=GREEN_FILL)
        input_cells_ws3.append(c)

    ws3.row_dimensions[row].height = 22
    row += 1

for c in input_cells_ws3:
    dv3.add(c)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 – Instrucciones
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("4 - Cómo usar 📋")
ws4.sheet_view.showGridLines = False
ws4.column_dimensions["A"].width = 4
ws4.column_dimensions["B"].width = 72

ws4.merge_cells("A1:B1")
apply_header(ws4["A1"], "📋  INSTRUCCIONES — CÓMO RESERVAR TU LUGAR", bg=BLUE_HEADER, size=13)
ws4.row_dimensions[1].height = 32

lines = [
    ("", ""),
    ("🎉", "¡Bienvenidos a la agenda familiar de los Mussali — Marzo 2026!"),
    ("", "La bebé está a punto de llegar. Esta agenda nos ayuda a organizarnos."),
    ("", ""),
    ("📱", "CÓMO FUNCIONA (muy fácil)"),
    ("", "  1. Toca la pestaña que te interesa (en la parte de abajo de la pantalla):"),
    ("", "       📚  1 - Emilio  →  para ayudar con la escuela de Emilio"),
    ("", "       🛁  2 - Baño Bebé  →  para bañar a la bebé ese día"),
    ("", "       🏠  3 - Visitas  →  para visitar la casa (3:00pm – 5:00pm)"),
    ("", "  2. Busca el día que te funcione."),
    ("", "  3. Toca la celda vacía → aparece un menú con los nombres → elige el tuyo."),
    ("", "  4. ¡Listo! Tu nombre queda guardado automáticamente."),
    ("", ""),
    ("📚", "ESCUELA DE EMILIO (pestaña 1)"),
    ("", "  • Llevar a Chabad:  Mar–Vie antes de las 9:00am"),
    ("", "  • Recoger Chabad → llevar a Village:  Mar–Vie ~11:45am"),
    ("", "  • Llevar a Village (solo Lunes):  Lunes a las 12:00pm"),
    ("", "  • Recoger de Village:  Lun–Vie ~3:00pm"),
    ("", "  ⚠️  Emilio debe llegar a casa antes de las 5:30pm (hora de baño)."),
    ("", ""),
    ("🛁", "BAÑO DE LA BEBÉ (pestaña 2)"),
    ("", "  • Solo 1 persona por día."),
    ("", "  • Primer lugar disponible = tuyo."),
    ("", ""),
    ("🏠", "VISITAS A CASA (pestaña 3)"),
    ("", "  • Cualquier día, de 3:00pm a 5:00pm."),
    ("", "  • Pueden venir hasta 3 grupos el mismo día."),
    ("", ""),
    ("❌", "REGLAS IMPORTANTES"),
    ("", "  • Si ya hay un nombre en una celda, ese lugar está tomado. Elige otro día."),
    ("", "  • No borres el nombre de alguien más."),
    ("", "  • Si necesitas cancelar o cambiar, avísale a Edu por WhatsApp."),
    ("", ""),
    ("💛", "¡Gracias por su amor y apoyo! Con cariño, Edu & familia 🍼"),
]

for i, (icon, text) in enumerate(lines, 2):
    ws4.row_dimensions[i].height = 18
    if icon:
        ci = ws4.cell(i, 1, value=icon)
        ci.font = Font(name="Segoe UI Emoji", size=11)
        ci.alignment = Alignment(horizontal="center", vertical="center")
    ct = ws4.cell(i, 2, value=text)
    if icon in ["📚", "🛁", "🏠", "❌", "💛", "📱"]:
        ct.font = Font(name="Calibri", bold=True, size=11, color="1F4E79")
        ct.fill = h_fill("EBF3FB")
    elif icon == "🎉":
        ct.font = Font(name="Calibri", bold=True, size=12, color="1F4E79")
    else:
        ct.font = Font(name="Calibri", size=10, color="2E2E2E")
    ct.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ── Save ──────────────────────────────────────────────────────────────────────
out = "/root/.openclaw/workspace/Agenda_Mussali_Marzo2026_v2.xlsx"
wb.save(out)
print(f"✅ Saved: {out}")
